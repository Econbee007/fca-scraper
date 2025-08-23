import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

URL = "https://fcainfoweb.nic.in/reports/report_menu_web.aspx"
OUT_FILE = "daily_prices_feb_apr2020.xlsx"


def scrape_one_date(driver, wait, target_date: str):
    """Navigate, enter date, wait for captcha solve, scrape table."""

    target_date_site = target_date.replace("-", "/")

    driver.get(URL)
    time.sleep(1)

    # 1) Select "Price Report"
    price_report_radio = wait.until(
        EC.element_to_be_clickable((By.ID, "ctl00_MainContent_Rbl_Rpt_type_0"))
    )
    price_report_radio.click()
    time.sleep(1)

    # 2) Choose "Daily Prices"
    daily_dropdown = wait.until(
        EC.presence_of_element_located((By.ID, "ctl00_MainContent_Ddl_Rpt_Option0"))
    )
    Select(daily_dropdown).select_by_visible_text("Daily Prices")
    time.sleep(1)

    # 3) Enter the date
    date_input = wait.until(
        EC.presence_of_element_located((By.ID, "ctl00_MainContent_Txt_FrmDate"))
    )
    driver.execute_script("arguments[0].value = '';", date_input)
    date_input.send_keys(target_date_site)
    print(f" Entered date: {target_date_site}")

    # 4) Wait for manual captcha solving + Get Data
    input("\n Please solve CAPTCHA in Chrome, click 'Get Data', "
          "and wait until the table loads. Then press ENTER here...")

    # 5) Scrape table
    table = wait.until(EC.presence_of_element_located((By.ID, "gv0")))
    rows = table.find_elements(By.TAG_NAME, "tr")

    data = []
    header = None
    for i, row in enumerate(rows):
        cells = [cell.text.strip() for cell in row.find_elements(By.TAG_NAME, "td")]
        if cells:
            if i == 0:
                header = ["Date"] + cells  # capture header row
            else:
                data.append([target_date_site] + cells)

    return pd.DataFrame(data, columns=header)


def save_incremental(df, out_file=OUT_FILE):
    """Append data to Excel incrementally (protects against crashes)."""
    try:
        book = load_workbook(out_file)
        with pd.ExcelWriter(out_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            startrow = writer.sheets["Sheet1"].max_row
            df.to_excel(writer, index=False, header=False, startrow=startrow)
    except FileNotFoundError:
        df.to_excel(out_file, index=False, header=True)


def get_existing_dates(out_file=OUT_FILE):
    """Return set of dates already present in Excel file (if it exists)."""
    try:
        # First try reading with headers
        df_existing = pd.read_excel(out_file)
        if "Date" in df_existing.columns:
            return set(df_existing["Date"].astype(str))
        else:
            # Fall back: assume first column is Date
            return set(df_existing.iloc[:, 0].astype(str))
    except FileNotFoundError:
        return set()


def sort_final_file(out_file=OUT_FILE):
    """Sort file by Date and overwrite clean."""
    try:
        df_final = pd.read_excel(out_file)

        # If no header, assign manually
        if "Date" not in df_final.columns:
            df_final.columns = ["Date"] + [f"Col{i}" for i in range(1, len(df_final.columns))]
        
        df_final["Date"] = pd.to_datetime(df_final["Date"], dayfirst=True, errors="coerce")
        df_final = df_final.sort_values("Date").reset_index(drop=True)
        df_final.to_excel(out_file, index=False)
        print(" Sorted dataset by Date and ensured headers are included.")
    except Exception as e:
        print(f" Could not sort final file: {e}")


if __name__ == "__main__":
    # Allow multiple dates in one input
    date_input = input(" Enter date(s) (dd-mm-yyyy), separated by commas: ").strip()
    date_list = [d.strip() for d in date_input.split(",")]

    # Load already-saved dates
    existing_dates = get_existing_dates(OUT_FILE)
    print(f" Found {len(existing_dates)} dates already in {OUT_FILE}")

    # Filter out duplicates
    missing_dates = [d for d in date_list if d not in existing_dates]

    if not missing_dates:
        print(" All requested dates are already saved. Nothing to do.")
    else:
        print(f" Dates to scrape: {missing_dates}")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.maximize_window()
        wait = WebDriverWait(driver, 40)

        for target_date in missing_dates:
            try:
                df = scrape_one_date(driver, wait, target_date)
                print(f" Scraped {len(df)} rows for {target_date}")
                save_incremental(df, OUT_FILE)
            except Exception as e:
                print(f" Failed for {target_date}: {e}")

        driver.quit()

        # Sort final dataset after appending
        sort_final_file(OUT_FILE)

        print(f"\n Finished appending + sorting data into {OUT_FILE}")
