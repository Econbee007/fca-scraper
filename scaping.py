import time
import pandas as pd
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

URL = "https://fcainfoweb.nic.in/reports/report_menu_web.aspx"
OUT_FILE = "daily_prices_feb_apr2020.xlsx"


def scrape_one_date(driver, wait, target_date: str):
    """Navigate, enter date, wait for captcha solve, scrape table."""

    # Convert format to dd/mm/yyyy (ASP.NET expects slashes)
    target_date_site = target_date.replace("-", "/")

    driver.get(URL)
    time.sleep(1)

    # 1) Select "Price Report"
    price_report_radio = wait.until(
        EC.element_to_be_clickable((By.ID, "ctl00_MainContent_Rbl_Rpt_type_0"))
    )
    price_report_radio.click()
    time.sleep(1)

    # 2) Choose "Daily Prices"
    daily_dropdown = wait.until(
        EC.presence_of_element_located((By.ID, "ctl00_MainContent_Ddl_Rpt_Option0"))
    )
    Select(daily_dropdown).select_by_visible_text("Daily Prices")
    time.sleep(1)

    # 3) Enter the date
    date_input = wait.until(
        EC.presence_of_element_located((By.ID, "ctl00_MainContent_Txt_FrmDate"))
    )
    driver.execute_script("arguments[0].value = '';", date_input)
    date_input.send_keys(target_date_site)
    print(f" Entered date: {target_date_site}")

    # 4) Wait for manual captcha solving + Get Data
    input("\n Please solve CAPTCHA in Chrome, click 'Get Data', "
          "and wait until the table loads. Then press ENTER here...")

    # 5) Scrape table gv0
    table = wait.until(EC.presence_of_element_located((By.ID, "gv0")))
    rows = table.find_elements(By.TAG_NAME, "tr")

    data = []
    for row in rows:
        cells = [cell.text.strip() for cell in row.find_elements(By.TAG_NAME, "td")]
        if cells:
            data.append([target_date_site] + cells)

    return pd.DataFrame(data)


def save_incremental(df, out_file=OUT_FILE):
    """Append data to Excel incrementally (protects against crashes)."""
    try:
        book = load_workbook(out_file)
        with pd.ExcelWriter(out_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            startrow = writer.sheets["Sheet1"].max_row
            df.to_excel(writer, index=False, header=False, startrow=startrow)
    except FileNotFoundError:
        df.to_excel(out_file, index=False, header=False)


def run_scraper(dates, out_file=OUT_FILE):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.maximize_window()
    wait = WebDriverWait(driver, 40)

    for d in dates:
        try:
            df = scrape_one_date(driver, wait, d)
            print(f" Scraped {len(df)} rows for {d}")
            save_incremental(df, out_file)   # save immediately
        except Exception as e:
            print(f" Failed for {d}: {e}")

    driver.quit()
    print(f"\n Scraping finished. Data saved to {out_file}")


if __name__ == "__main__":
    # Auto-generate all dates from Feb 1 to Apr 30, 2020
    start = datetime.strptime("01-02-2020", "%d-%m-%Y")
    end = datetime.strptime("30-04-2020", "%d-%m-%Y")

    date_list = [(start + timedelta(days=i)).strftime("%d-%m-%Y")
                 for i in range((end - start).days + 1)]
    run_scraper(date_list)
